"""واردکردنِ لیستِ قیمتِ تامین‌کننده از اکسل/PDF/عکس -- طبقِ درخواستِ صریح:
هر کالا می‌تواند چند «کدِ تامین‌کننده» داشته باشد (کدِ خودِ کالا نزدِ آن
تامین‌کننده، که با کدِ داخلیِ ما فرق دارد) تا ردیف‌هایِ فایلِ قیمتِ او
خودکار به کالایِ داخلی متصل شوند. سپس رویِ قیمتِ تامین‌کننده چند ستونِ
افزایشیِ درصدی/مبلغی اعمال و نتیجه در یک فهرستِ قیمتِ موجود ثبت می‌شود.

فازِ اول فقط اکسل و PDFِ متنی را پشتیبانی می‌کرد. **فازِ دوم**: پشتیبانیِ
عکس (jpg/png) و PDFِ اسکن‌شده (بدونِ لایه‌یِ متن) با OCR -- از Tesseract
OCR (مجوزِ Apache-2.0، سازگار با توزیعِ تجاری/بسته) از طریقِ pytesseract
استفاده شده. برخلافِ استخراجِ PDFِ متنی که تقریباً همیشه دقیق است، OCR
ذاتاً خطاپذیر است (خصوصاً وقتی کد/توضیح به خطِ فارسی باشد کنارِ کدهایِ
لاتین) -- به همین دلیل: (۱) در UI همیشه یک هشدارِ «نیازمندِ بازبینیِ
دستی» نشان داده می‌شود، (۲) همان مسیرِ «🔗 اتصال»ِ فازِ اول برایِ
ردیف‌هایی که کدشان غلط خوانده شده هم کار می‌کند (چون آن ردیف صرفاً
تطبیق‌نیافته دیده می‌شود، نه اینکه به کالایِ اشتباه بچسبد). آزمایش نشان
داد OCR با lang="eng" برایِ ستون‌هایِ کد/قیمت (که تقریباً همیشه
لاتین/رقمی‌اند) دقیقِ بسیار بهتری از حالتِ ترکیبیِ "fas+eng" دارد --
چون توضیحِ فارسی برایِ تطبیق استفاده نمی‌شود، پیش‌فرض را رویِ "eng"
گذاشتیم و "fas+eng" را به‌عنوانِ گزینه برایِ کدهایِ فارسی/مختلط نگه
داشتیم."""

from __future__ import annotations

import decimal
import re
import statistics
from dataclasses import dataclass

import openpyxl
import pdfplumber
from sqlalchemy import select

from peecha import numerals
from peecha.db.base import new_session
from peecha.db.models.commercial import SupplierPriceImportTemplate
from peecha.db.models.inventory import Item, ItemSupplierCode

_ZERO = decimal.Decimal("0")
_Q2 = decimal.Decimal("0.01")

# طبقِ باگِ گزارش‌شده («کد از فایل و کدِ ثبت‌شده در تنظیماتِ کالا هردو
# ۱۰۰۱ دیده می‌شوند ولی تطبیق نمی‌خورند»): کاراکترهایِ نامرئی (نیم‌فاصله،
# نشانه‌هایِ جهت، فاصله‌یِ بدونِ‌شکست، BOM) اغلب هنگامِ کپی از اکسل/PDF/
# سایتِ تامین‌کننده وارد متن می‌شوند و با چشم دیده نمی‌شوند، ولی برایِ
# مقایسه‌یِ رشته‌ای دو مقدار را «متفاوت» می‌کنند. str.strip() فقط انتهایِ
# رشته را پاک می‌کند، نه وسطِ آن را -- پس این‌ها را از کلِ رشته حذف می‌کنیم.
_INVISIBLE_CHARS_RE = re.compile(
    "[​‌‍‎‏﻿\xa0]"
)  # ZWSP, ZWNJ (نیم‌فاصله), ZWJ, LRM, RLM, BOM, NBSP


def _strip_invisible_chars(text: str) -> str:
    return _INVISIBLE_CHARS_RE.sub("", text)


def _normalize_code(code: str) -> str:
    text = _strip_invisible_chars(numerals.to_ascii_digits(code)).strip().upper()
    # کدهایِ کاملاً عددی: اکسل معمولاً صفرهایِ ابتدایی را در نوعِ «عدد» از
    # دست می‌دهد (۰۱۰۰۱ می‌شود ۱۰۰۱)، درحالی‌که کدِ ثبت‌شده در تنظیماتِ
    # کالا ممکن است با صفرِ ابتدایی تایپ شده باشد -- برایِ جلوگیری از
    # عدم‌تطبیقِ کاذب، صفرهایِ ابتدایی را نادیده می‌گیریم.
    if text.isdigit() and len(text) > 1:
        text = text.lstrip("0") or "0"
    return text


def _normalize_name(name: str) -> str:
    """طبقِ درخواستِ صریح («بعضی تامین‌کننده‌ها فقط نامِ کالا دارند»):
    نام‌ها معمولاً در چاپ‌هایِ مختلف با فاصله/نیم‌فاصله متفاوت‌اند --
    این تابع همه را به یک شکلِ یکسان می‌رساند تا تطبیقِ دقیق ممکن شود."""
    text = _strip_invisible_chars(numerals.to_ascii_digits(name)).strip().casefold()
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_by_type(value: str, value_type: str) -> str:
    return _normalize_code(value) if value_type == "CODE" else _normalize_name(value)


# ---------------------------------------------------------------------
# کدها/نام‌هایِ تامین‌کننده برایِ هر کالا -- طبقِ درخواستِ صریح، تشخیص
# باید «ترکیبی» باشد: هم کد و هم نام، هرکدام موجود باشد کافی است.
# ---------------------------------------------------------------------
@dataclass
class ItemSupplierCodeRow:
    item_supplier_code_id: int
    item_id: int
    supplier_detail_account_id: int | None
    value_type: str  # CODE | NAME
    supplier_code: str


def list_item_supplier_codes(item_id: int) -> list[ItemSupplierCodeRow]:
    with new_session() as session:
        rows = session.scalars(
            select(ItemSupplierCode).where(ItemSupplierCode.item_id == item_id)
            .order_by(ItemSupplierCode.value_type, ItemSupplierCode.item_supplier_code_id)
        ).all()
        return [
            ItemSupplierCodeRow(r.item_supplier_code_id, r.item_id, r.supplier_detail_account_id, r.value_type, r.supplier_code)
            for r in rows
        ]


def add_item_supplier_code(
    item_id: int, supplier_code: str, supplier_detail_account_id: int | None = None, value_type: str = "CODE",
) -> int:
    if value_type not in ("CODE", "NAME"):
        raise ValueError("نوعِ مقدار باید کد یا نام باشد.")
    supplier_code = supplier_code.strip()
    if not supplier_code:
        raise ValueError("مقدار نمی‌تواند خالی باشد.")
    normalized = _normalize_by_type(supplier_code, value_type)
    with new_session() as session:
        existing = session.scalar(
            select(ItemSupplierCode).where(
                ItemSupplierCode.item_id == item_id,
                ItemSupplierCode.supplier_detail_account_id == supplier_detail_account_id,
                ItemSupplierCode.value_type == value_type,
                ItemSupplierCode.normalized_code == normalized,
            )
        )
        if existing is not None:
            raise ValueError("این مقدار قبلاً برایِ همین کالا (نزدِ همین تامین‌کننده) ثبت شده است.")
        row = ItemSupplierCode(
            item_id=item_id, supplier_detail_account_id=supplier_detail_account_id, value_type=value_type,
            supplier_code=supplier_code, normalized_code=normalized,
        )
        session.add(row)
        session.commit()
        return row.item_supplier_code_id


def update_item_supplier_code(
    item_supplier_code_id: int, supplier_code: str, supplier_detail_account_id: int | None, value_type: str,
) -> None:
    if value_type not in ("CODE", "NAME"):
        raise ValueError("نوعِ مقدار باید کد یا نام باشد.")
    supplier_code = supplier_code.strip()
    if not supplier_code:
        raise ValueError("مقدار نمی‌تواند خالی باشد.")
    normalized = _normalize_by_type(supplier_code, value_type)
    with new_session() as session:
        row = session.get(ItemSupplierCode, item_supplier_code_id)
        if row is None:
            raise ValueError("این مقدار یافت نشد.")
        duplicate = session.scalar(
            select(ItemSupplierCode).where(
                ItemSupplierCode.item_id == row.item_id,
                ItemSupplierCode.supplier_detail_account_id == supplier_detail_account_id,
                ItemSupplierCode.value_type == value_type,
                ItemSupplierCode.normalized_code == normalized,
                ItemSupplierCode.item_supplier_code_id != item_supplier_code_id,
            )
        )
        if duplicate is not None:
            raise ValueError("این مقدار قبلاً برایِ همین کالا (نزدِ همین تامین‌کننده) ثبت شده است.")
        row.supplier_code = supplier_code
        row.supplier_detail_account_id = supplier_detail_account_id
        row.value_type = value_type
        row.normalized_code = normalized
        session.commit()


def delete_item_supplier_code(item_supplier_code_id: int) -> None:
    with new_session() as session:
        row = session.get(ItemSupplierCode, item_supplier_code_id)
        if row is None:
            raise ValueError("این مقدار یافت نشد.")
        session.delete(row)
        session.commit()


def find_item_by_supplier_reference(company_id: int, raw_text: str, supplier_detail_account_id: int | None) -> int | None:
    """تشخیصِ ترکیبی: اول کد را امتحان می‌کند (اول نزدِ همین تامین‌کننده،
    بعد کدِ عمومی)؛ اگر چیزی پیدا نشد، همان متن را به‌عنوانِ «نام» امتحان
    می‌کند (باز هم اول نزدِ همین تامین‌کننده، بعد نامِ عمومی) -- چون
    بعضی تامین‌کننده‌ها به‌جایِ کد، فقط نامِ کالا را در فایل می‌نویسند."""
    if not raw_text:
        return None
    with new_session() as session:
        for value_type in ("CODE", "NAME"):
            normalized = _normalize_by_type(raw_text, value_type)
            if not normalized:
                continue
            if supplier_detail_account_id is not None:
                row = session.scalar(
                    select(ItemSupplierCode)
                    .join(Item, Item.item_id == ItemSupplierCode.item_id)
                    .where(
                        Item.company_id == company_id,
                        ItemSupplierCode.supplier_detail_account_id == supplier_detail_account_id,
                        ItemSupplierCode.value_type == value_type,
                        ItemSupplierCode.normalized_code == normalized,
                    )
                )
                if row is not None:
                    return row.item_id
            row = session.scalar(
                select(ItemSupplierCode)
                .join(Item, Item.item_id == ItemSupplierCode.item_id)
                .where(
                    Item.company_id == company_id,
                    ItemSupplierCode.supplier_detail_account_id.is_(None),
                    ItemSupplierCode.value_type == value_type,
                    ItemSupplierCode.normalized_code == normalized,
                )
            )
            if row is not None:
                return row.item_id
    return None


# ---------------------------------------------------------------------
# قالبِ ذخیره‌شده‌یِ ستون‌بندیِ فایلِ هر تامین‌کننده
# ---------------------------------------------------------------------
@dataclass
class ImportTemplateRow:
    template_id: int
    supplier_detail_account_id: int
    code_column_index: int
    price_column_index: int
    header_row_index: int
    sheet_name: str | None


def get_import_template(company_id: int, supplier_detail_account_id: int) -> ImportTemplateRow | None:
    with new_session() as session:
        row = session.scalar(
            select(SupplierPriceImportTemplate).where(
                SupplierPriceImportTemplate.company_id == company_id,
                SupplierPriceImportTemplate.supplier_detail_account_id == supplier_detail_account_id,
            )
        )
        if row is None:
            return None
        return ImportTemplateRow(
            row.template_id, row.supplier_detail_account_id, row.code_column_index, row.price_column_index,
            row.header_row_index, row.sheet_name,
        )


def save_import_template(
    company_id: int, supplier_detail_account_id: int, code_column_index: int, price_column_index: int,
    header_row_index: int, sheet_name: str | None = None,
) -> None:
    with new_session() as session:
        row = session.scalar(
            select(SupplierPriceImportTemplate).where(
                SupplierPriceImportTemplate.company_id == company_id,
                SupplierPriceImportTemplate.supplier_detail_account_id == supplier_detail_account_id,
            )
        )
        if row is None:
            row = SupplierPriceImportTemplate(company_id=company_id, supplier_detail_account_id=supplier_detail_account_id)
            session.add(row)
        row.code_column_index = code_column_index
        row.price_column_index = price_column_index
        row.header_row_index = header_row_index
        row.sheet_name = sheet_name
        session.commit()


# ---------------------------------------------------------------------
# استخراجِ جدولِ خام از فایل -- طبقِ توافق: فقط اکسل و PDFِ متنی
# ---------------------------------------------------------------------
def list_excel_sheet_names(file_path: str) -> list[str]:
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def extract_excel_grid(file_path: str, sheet_name: str | None = None) -> list[list[str]]:
    workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
        return [["" if cell is None else str(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def is_pdf_text_based(file_path: str) -> bool:
    """اگر حتی یک صفحه لایهٔ متنِ واقعی داشته باشد True -- یعنی PDF متنی
    است و extract_pdf_grid کافی است. اگر False باشد یعنی این PDF یک
    اسکنِ عکسی است و باید با extract_pdf_grid_ocr خوانده شود."""
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            if page.extract_words():
                return True
    return False


def is_ocr_available() -> bool:
    """آیا Tesseract OCR رویِ این سیستم نصب است. pytesseract فقط یک پوستهٔ
    نازک است -- خودِ برنامهٔ Tesseract باید جداگانه رویِ سیستم نصب شده
    باشد (این یک وابستگیِ سیستمی است، نه یک بستهٔ pip)."""
    try:
        import pytesseract

        pytesseract.get_tesseract_version()
        return True
    except Exception:
        return False


_OCR_MIN_COLUMN_GAP_PX = 15.0
_OCR_COLUMN_GAP_RATIO = 1.5
_OCR_PSM = 6  # «یک بلوکِ یکنواختِ متن» -- برایِ متنِ جدولی/ردیفی به‌مراتب دقیق‌تر از حالتِ خودکارِ پیش‌فرض


def _extract_grid_from_ocr_image(image, lang: str) -> list[list[str]]:
    """معادلِ _extract_grid_from_words ولی رویِ خروجیِ OCR: کلمه‌هایِ
    هم‌خط (بر اساسِ block/paragraph/line) را کنارِ هم می‌گذارد و هرجا
    فاصله‌یِ افقی از یک آستانه (نسبت به قدِ متوسطِ کلمه‌ها -- نه یک عددِ
    ثابتِ پیکسلی، چون وضوحِ تصویر متغیر است) بیشتر شود، ستونِ تازه شروع
    می‌کند."""
    import pytesseract

    data = pytesseract.image_to_data(image, lang=lang, config=f"--psm {_OCR_PSM}", output_type=pytesseract.Output.DICT)
    words: list[dict] = []
    for i in range(len(data["text"])):
        text = data["text"][i].strip()
        if not text:
            continue
        try:
            conf = float(data["conf"][i])
        except (TypeError, ValueError):
            conf = -1.0
        if conf < 0:
            continue
        words.append({
            "text": text, "left": data["left"][i], "width": data["width"][i], "height": data["height"][i],
            "line_key": (data["block_num"][i], data["par_num"][i], data["line_num"][i]),
        })
    if not words:
        return []
    median_height = statistics.median(w["height"] for w in words)
    gap_threshold = max(_OCR_MIN_COLUMN_GAP_PX, median_height * _OCR_COLUMN_GAP_RATIO)
    rows: dict[tuple, list[dict]] = {}
    for w in words:
        rows.setdefault(w["line_key"], []).append(w)
    grid: list[list[str]] = []
    for key in sorted(rows.keys()):
        row_words = sorted(rows[key], key=lambda w: w["left"])
        cells: list[str] = []
        current_cell = [row_words[0]["text"]]
        prev_right = row_words[0]["left"] + row_words[0]["width"]
        for w in row_words[1:]:
            if w["left"] - prev_right > gap_threshold:
                cells.append(" ".join(current_cell))
                current_cell = [w["text"]]
            else:
                current_cell.append(w["text"])
            prev_right = w["left"] + w["width"]
        cells.append(" ".join(current_cell))
        grid.append(cells)
    return grid


def extract_image_grid(file_path: str, lang: str = "eng") -> list[list[str]]:
    from PIL import Image

    with Image.open(file_path) as image:
        return _extract_grid_from_ocr_image(image, lang)


def extract_pdf_grid_ocr(file_path: str, lang: str = "eng", dpi: int = 300) -> list[list[str]]:
    """رندرِ هر صفحهٔ PDFِ اسکن‌شده به تصویر و سپس OCR -- برایِ زمانی که
    is_pdf_text_based همان فایل False برگردانده باشد."""
    import pypdfium2 as pdfium

    pdf = pdfium.PdfDocument(file_path)
    try:
        scale = dpi / 72
        grid: list[list[str]] = []
        for page in pdf:
            bitmap = page.render(scale=scale)
            grid.extend(_extract_grid_from_ocr_image(bitmap.to_pil(), lang))
        return grid
    finally:
        pdf.close()


def extract_pdf_grid(file_path: str) -> list[list[str]]:
    """اول با تشخیصِ جدولِ بومیِ pdfplumber (بهترین دقت، وقتی PDF واقعاً
    خط‌کشیِ جدول دارد)؛ اگر چیزی پیدا نشد (رایج در فاکتور/لیست‌هایِ
    بدونِ خط‌کشیِ صریح)، به استخراجِ متنِ خام برمی‌گردد و ستون‌ها را از
    رویِ فاصله‌هایِ متوالی (رایج‌ترین الگویِ تراز-چپ/راستِ جدولی) حدس
    می‌زند. اگر PDF یک اسکنِ عکسی باشد (بدونِ لایه‌یِ متن)، هردو راه چیزی
    برنمی‌گردانند -- این حالت را UI با is_pdf_text_based از قبل تشخیص
    می‌دهد و به‌جایِ این تابع، extract_pdf_grid_ocr را صدا می‌زند."""
    grid: list[list[str]] = []
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            for table in page.extract_tables():
                for row in table:
                    grid.append(["" if cell is None else str(cell).strip() for cell in row])
    if grid:
        return grid
    # طبقِ باگِ کشف‌شده حینِ آزمایش: extract_text() فاصله‌هایِ چندتاییِ
    # واقعی را به یک فاصله تبدیل می‌کند (بازسازیِ طبیعیِ متن، نه حفظِ
    # چیدمانِ ستونی) -- پس extract_words() با مختصاتِ x استفاده می‌شود:
    # کلماتی که رویِ همان خط‌اند ولی فاصله‌یِ افقیِ زیاد (فاصله‌یِ ستون،
    # نه فاصله‌یِ بینِ‌کلمه‌ایِ معمولی) دارند، به دو ستونِ جدا تقسیم می‌شوند.
    with pdfplumber.open(file_path) as pdf:
        for page in pdf.pages:
            grid.extend(_extract_grid_from_words(page))
    return grid


def _extract_grid_from_words(page, column_gap_threshold: float = 10.0) -> list[list[str]]:
    words = page.extract_words()
    if not words:
        return []
    rows: dict[int, list[dict]] = {}
    for w in words:
        key = round(w["top"] / 3)
        rows.setdefault(key, []).append(w)
    grid: list[list[str]] = []
    for key in sorted(rows.keys()):
        row_words = sorted(rows[key], key=lambda w: w["x0"])
        cells: list[str] = []
        current_cell = [row_words[0]["text"]]
        prev_x1 = row_words[0]["x1"]
        for w in row_words[1:]:
            if w["x0"] - prev_x1 > column_gap_threshold:
                cells.append(" ".join(current_cell))
                current_cell = [w["text"]]
            else:
                current_cell.append(w["text"])
            prev_x1 = w["x1"]
        cells.append(" ".join(current_cell))
        grid.append(cells)
    return grid


def parse_price(text: str) -> decimal.Decimal | None:
    if not text:
        return None
    cleaned = numerals.to_ascii_digits(str(text)).replace(",", "").replace("٬", "").strip()
    cleaned = re.sub(r"[^\d.\-]", "", cleaned)
    if not cleaned or cleaned in ("-", "."):
        return None
    try:
        return decimal.Decimal(cleaned)
    except decimal.InvalidOperation:
        return None


# ---------------------------------------------------------------------
# تطبیقِ ردیف‌ها با کالاهایِ داخلی
# ---------------------------------------------------------------------
@dataclass
class MatchedPriceRow:
    row_no: int
    raw_code: str
    supplier_price: decimal.Decimal | None
    item_id: int | None
    item_label: str


def match_grid_rows(
    company_id: int, grid: list[list[str]], code_column: int, price_column: int, header_row_index: int,
    supplier_detail_account_id: int | None,
) -> list[MatchedPriceRow]:
    from peecha.services import inventory_catalog as catalog_service

    items_by_id = {i.item_id: i for i in catalog_service.list_items(company_id)}
    results: list[MatchedPriceRow] = []
    for row_no, row in enumerate(grid):
        if row_no <= header_row_index:
            continue
        if code_column >= len(row) or price_column >= len(row):
            continue
        raw_code = (row[code_column] or "").strip()
        raw_price_text = (row[price_column] or "").strip()
        if not raw_code and not raw_price_text:
            continue
        price = parse_price(raw_price_text)
        item_id = find_item_by_supplier_reference(company_id, raw_code, supplier_detail_account_id) if raw_code else None
        item = items_by_id.get(item_id) if item_id is not None else None
        item_label = f"{item.code} — {item.name or ''}" if item else ""
        results.append(MatchedPriceRow(row_no=row_no, raw_code=raw_code, supplier_price=price, item_id=item_id, item_label=item_label))
    return results


# ---------------------------------------------------------------------
# اعمالِ ستون‌هایِ افزایشیِ درصدی/مبلغی
# ---------------------------------------------------------------------
@dataclass
class PriceAdjustmentStep:
    kind: str  # PERCENT | AMOUNT
    value: decimal.Decimal
    label: str = ""


def apply_adjustments(base_price: decimal.Decimal, steps: list[PriceAdjustmentStep]) -> decimal.Decimal:
    price = base_price
    for step in steps:
        if step.kind == "PERCENT":
            price = price + (price * step.value / decimal.Decimal(100))
        elif step.kind == "AMOUNT":
            price = price + step.value
        else:
            raise ValueError(f"نوعِ ستونِ افزایشی نامعتبر است: {step.kind}")
    return price.quantize(_Q2, rounding=decimal.ROUND_HALF_UP)
