"""زیرساختِ مشترکِ خروجیِ گزارش‌ها — چاپ/PDF با QtPrintSupport (بدونِ
وابستگیِ بیرونی) و Excel با openpyxl. هر صفحه‌یِ گزارش (reports_common.py و
زیرکلاس‌هایش) همین سه تابع را با دیتایِ جدولِ فعلی‌اش صدا می‌زند.

طبقِ گزارش‌هایِ صریحِ کاربر، چند مشکلِ واقعی در خروجیِ چاپ/PDF پیدا و رفع شد:

۱. ترتیبِ ستون‌ها برعکس بود — ترتیبِ سلول‌ها در خودِ HTML معکوس نوشته
   می‌شود (نکته‌ی فنی: QTextDocument با dir="rtl" ستونِ *اول*ِ HTML را
   چپ‌ترین رسم می‌کند، برعکسِ قراردادِ راست‌به‌چپ).

۲. حاشیه‌یِ چاپ به‌صراحت به ۶ میلی‌متر تنظیم می‌شود.

۳. فونتِ چاپ از فونتِ واقعیِ برنامه (get_font_family) استفاده می‌کند،
   با اندازه‌یِ پیش‌فرضِ ۹pt (کوچک‌تر از حالتِ قبلی).

۴. **ریشه‌یِ واقعیِ باگِ «صفحه‌های نصفه/خالی» پیدا شد**: نسخه‌هایِ قبلی
   گزارش را به چند جدولِ جدا با CSS `page-break-after: always` بینِ آن‌ها
   می‌شکستند — پشتیبانیِ QTextDocument از page-break-* رویِ جدول‌هایِ
   چندبخشی قابلِ‌اتکا نیست و صفحه‌یِ خالیِ اضافه می‌ساخت.

۵. **هدر/فوترِ تکرارشونده + شماره‌یِ صفحه رویِ همه‌ی صفحات** (طبقِ
   درخواستِ صریح): چون `QTextDocument.print_` به‌تنهایی چنین چیزی نمی‌دهد
   (و شکستِ دستی هم باگ‌زا بود)، چاپ حالا به‌صورتِ دستی و صفحه‌به‌صفحه با
   `QPainter` انجام می‌شود (`_paint_report`): یک نوارِ هدرِ ثابت (که عیناً
   رویِ هر صفحه تکرار می‌شود) بالا، یک تکه از جدولِ محتوا (که خودِ Qt
   بومی و درست بینِ صفحات می‌شکند، بدونِ CSSِ دستی) در وسط، و یک نوارِ
   فوتر (متنِ کاربر + «صفحه‌یِ X از Y») پایینِ هر صفحه. ردیف‌هایِ «جمعِ
   این صفحه» هم به‌عنوانِ ردیفِ عادیِ همان جدولِ محتوا، در فاصله‌هایِ
   تخمین‌زده‌شده (یا دستی‌تنظیم‌شده) درج می‌شوند.

۶. **فرمِ «تنظیماتِ چاپ»** پیش از هر چاپ/PDF/اکسل باز می‌شود: اندازه‌یِ
   فونت، عرضِ هر ستون (٪)، تعدادِ ردیف در هر صفحه (یا خودکار)، و متنِ
   *کاملِ* هدر/فوتر (چندخطی، با متنِ پیش‌فرضِ همان نام‌شرکت/عنوان/تاریخ
   به‌عنوانِ نقطه‌یِ شروعِ قابلِ‌ویرایش) — این تنظیمات با QSettings رویِ
   دیسک (بر اساسِ عنوانِ هر گزارش) ذخیره می‌شوند و در اجراهایِ بعدی هم
   می‌مانند."""

from __future__ import annotations

import decimal
from dataclasses import dataclass

from PySide6.QtCore import QMarginsF, QRectF, QSettings, QSizeF
from PySide6.QtGui import QAbstractTextDocumentLayout, QPageLayout, QPainter, QTextDocument
from PySide6.QtPrintSupport import QPrinter, QPrintPreviewDialog
from PySide6.QtWidgets import (
    QCheckBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QLabel,
    QMessageBox,
    QScrollArea,
    QSpinBox,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

from peecha import numerals
from peecha.ui.main import get_font_family

_PAGE_MARGIN_MM = 6.0
_DEFAULT_FONT_SIZE_PT = 9.0
# طبقِ عنوانِ ستون تشخیص داده می‌شود که «جمعِ صفحه» برایش معنا دارد —
# نه فقط قابلِ‌پارس‌بودنِ عدد، وگرنه ستون‌هایی مثلِ «شماره‌یِ سند» هم
# اشتباهی جمع می‌شدند.
_AMOUNT_HEADER_KEYWORDS = ("بدهکار", "بستانکار", "بد)", "بس)", "مانده", "مبلغ")


@dataclass
class PrintOptions:
    font_size_pt: float = _DEFAULT_FONT_SIZE_PT
    # درصدِ عرضِ هر ستون، هم‌ترتیب با headers (نه معکوس‌شده) — None یعنی
    # عرضِ خودکار/مساوی.
    column_widths: list[float] | None = None
    # اگر خالی باشد، هدرِ خودکار (نامِ شرکت+عنوان+تاریخ+فیلترها) ساخته
    # می‌شود؛ اگر پر باشد، *کاملاً جایگزینِ* آن می‌شود (چندخطی، آزاد).
    header_text: str = ""
    # زیرِ جدول (پس از ردیفِ جمعِ کل) اضافه می‌شود — مثلاً محلِ امضا.
    footer_text: str = ""
    # ۰ یعنی خودکار (تخمینِ تعدادِ ردیفی که در یک صفحه جا می‌شود)؛ عددِ
    # مثبت یعنی کاربر خودش فاصله‌یِ ردیف‌هایِ «جمعِ این صفحه» را تعیین کرده.
    rows_per_page: int = 0
    # طبقِ درخواستِ صریح: نمایشِ خطوطِ جدول (کادرِ دورِ هر سلول) قابلِ‌
    # خاموش‌کردن باشد.
    show_gridlines: bool = True


def _settings_key(title: str) -> str:
    return f"print_options/{title}"


def load_print_options(title: str) -> PrintOptions | None:
    """طبقِ درخواستِ صریح («تنظیماتِ هر گزارش ذخیره بشه، هر بار تغییر
    نکنه»): تنظیماتِ چاپِ هر گزارش با QSettings (رویِ دیسک، مستقل از
    اجراهایِ بعدیِ برنامه) ذخیره/بازخوانی می‌شود — کلید بر اساسِ عنوانِ
    گزارش، پس هر گزارش تنظیماتِ خودش را جدا نگه می‌دارد."""
    settings = QSettings("Peecha", "PeechaERP")
    key = _settings_key(title)
    if not settings.contains(f"{key}/font_size_pt"):
        return None
    widths_str = settings.value(f"{key}/column_widths", "", type=str)
    column_widths = [float(x) for x in widths_str.split(",") if x] if widths_str else None
    return PrintOptions(
        font_size_pt=float(settings.value(f"{key}/font_size_pt", _DEFAULT_FONT_SIZE_PT)),
        column_widths=column_widths,
        header_text=settings.value(f"{key}/header_text", "", type=str),
        footer_text=settings.value(f"{key}/footer_text", "", type=str),
        rows_per_page=int(settings.value(f"{key}/rows_per_page", 0)),
        show_gridlines=bool(settings.value(f"{key}/show_gridlines", True, type=bool)),
    )


def save_print_options(title: str, options: PrintOptions) -> None:
    settings = QSettings("Peecha", "PeechaERP")
    key = _settings_key(title)
    settings.setValue(f"{key}/font_size_pt", options.font_size_pt)
    settings.setValue(
        f"{key}/column_widths", ",".join(str(w) for w in options.column_widths) if options.column_widths else ""
    )
    settings.setValue(f"{key}/header_text", options.header_text)
    settings.setValue(f"{key}/footer_text", options.footer_text)
    settings.setValue(f"{key}/rows_per_page", options.rows_per_page)
    settings.setValue(f"{key}/show_gridlines", options.show_gridlines)


def _escape(value: object) -> str:
    return str(value).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _print_font_family() -> str:
    try:
        return get_font_family()
    except Exception:
        return "Tahoma"


def _amount_column_indices(headers: list[str], rows: list[list]) -> list[int]:
    indices = []
    for i, header in enumerate(headers):
        if not any(kw in header for kw in _AMOUNT_HEADER_KEYWORDS):
            continue
        parseable = True
        for row in rows:
            cell = str(row[i]).strip() if i < len(row) else ""
            if not cell:
                continue
            try:
                numerals.parse_decimal(cell)
            except ValueError:
                parseable = False
                break
        if parseable:
            indices.append(i)
    return indices


def _sum_amount_cells(rows: list[list], col_index: int) -> decimal.Decimal:
    total = decimal.Decimal(0)
    for row in rows:
        cell = str(row[col_index]).strip() if col_index < len(row) else ""
        if cell:
            total += numerals.parse_decimal(cell)
    return total


def _subtotal_row(headers: list[str], amount_col_indices: list[int], chunk_rows: list[list]) -> list:
    totals: list[str] = ["" for _ in headers]
    totals[0] = "جمعِ این صفحه"
    for col in amount_col_indices:
        totals[col] = numerals.format_amount(_sum_amount_cells(chunk_rows, col))
    return totals


def _default_header_plain_text(title: str, company_name: str, report_date: str, filters: list[tuple[str, str]] | None) -> str:
    lines = []
    if company_name:
        lines.append(company_name)
    lines.append(title)
    meta_parts = []
    if report_date:
        meta_parts.append(f"تاریخِ گزارش: {report_date}")
    for label, value in filters or []:
        meta_parts.append(f"{label}: {value}")
    if meta_parts:
        lines.append(" | ".join(meta_parts))
    return "\n".join(lines)


def _auto_header_html(title: str, company_name: str, report_date: str, filters: list[tuple[str, str]] | None) -> str:
    header_lines = ""
    if company_name:
        header_lines += f'<div style="font-size:12pt; font-weight:bold;">{_escape(company_name)}</div>'
    header_lines += f'<h3 style="margin:4px 0;">{_escape(title)}</h3>'
    meta_parts = []
    if report_date:
        meta_parts.append(f"تاریخِ گزارش: {_escape(report_date)}")
    for label, value in filters or []:
        meta_parts.append(f"{_escape(label)}: {_escape(value)}")
    if meta_parts:
        header_lines += (
            '<div style="font-size:9pt; color:#444;">' + " &nbsp;|&nbsp; ".join(meta_parts) + "</div>"
        )
    return header_lines


def _multiline_html(text: str) -> str:
    lines = text.splitlines() or [""]
    return "".join(f"<div>{_escape(line) if line.strip() else '&nbsp;'}</div>" for line in lines)


def _resolve_header_html(options: PrintOptions, title: str, company_name: str, report_date: str, filters: list[tuple[str, str]] | None) -> str:
    if options.header_text.strip():
        return _multiline_html(options.header_text)
    return _auto_header_html(title, company_name, report_date, filters)


def _colgroup_html(column_widths: list[float] | None, num_cols: int) -> str:
    if not column_widths or len(column_widths) != num_cols:
        return ""
    reversed_widths = list(reversed(column_widths))
    cols = "".join(f'<col style="width:{w:.2f}%">' for w in reversed_widths)
    return f"<colgroup>{cols}</colgroup>"


# طبقِ دو گزارشِ متناقض‌نما که هردو با آزمایشِ مستقیم بررسی شدند:
# ابتدا گمان رفت مشکل این است که کادرِ کلِ <table> در لبه‌ی صفحه بسته
# نمی‌شود، پس حاشیه به CSSِ هر سلول (border:...pt) منتقل شد. اما با
# رندرِ واقعیِ QTextDocument به QImage و شمارشِ پیکسل ثابت شد که Qt
# اصلاً به CSSِ `border` رویِ <td> بی‌توجه است — هیچ خطی رسم نمی‌شود،
# صرفِ‌نظر از مقدار. در همان آزمایش، attributeِ قدیمیِ HTMLِ `border="1"`
# رویِ خودِ <table> واقعاً و همیشه رسم می‌شود (این مسیرِ رندرِ بومیِ
# QTextTable است، نه یک قاعده‌ی CSSِ عمومی). پس به همان attribute
# برمی‌گردیم؛ برایِ تیکِ «نمایشِ خطوط»، مقدارِ attribute بینِ ۱ و ۰
# عوض می‌شود.
def _table_border_attr(show_gridlines: bool) -> str:
    return "1" if show_gridlines else "0"


def _rows_html(rows: list[list]) -> str:
    return "".join(
        "<tr>" + "".join(f"<td>{_escape(c)}</td>" for c in reversed(row)) + "</tr>" for row in rows
    )


def _bold_row_html(cells: list) -> str:
    return "<tr>" + "".join(f"<td><b>{_escape(c)}</b></td>" for c in reversed(cells)) + "</tr>"


def _content_table_html(colgroup_html: str, head_cells: str, body_rows_html: str, extra_row_html: str, show_gridlines: bool) -> str:
    return f"""
    <table border="{_table_border_attr(show_gridlines)}" cellspacing="0" cellpadding="3" width="100%" style="border-collapse: collapse; table-layout: fixed;">
    {colgroup_html}
    <thead><tr>{head_cells}</tr></thead>
    <tbody>{body_rows_html}{extra_row_html}</tbody>
    </table>
    """


def _wrap_document(body_html: str, font_family: str, font_size_pt: float) -> str:
    return f"""
    <html dir="rtl"><head><meta charset="utf-8"></head>
    <body style="font-family: '{font_family}', Tahoma, sans-serif; font-size: {font_size_pt}pt;">
    {body_html}
    </body></html>
    """


def _fits_one_page(page_size: QSizeF, table_html: str, font_family: str, font_size_pt: float) -> bool:
    doc = QTextDocument()
    doc.setHtml(_wrap_document(table_html, font_family, font_size_pt))
    doc.setPageSize(page_size)
    return doc.pageCount() <= 1


def _estimate_rows_per_page(
    rows: list[list],
    headers: list[str],
    amount_col_indices: list[int],
    *,
    page_size: QSizeF,
    colgroup_html: str,
    head_cells: str,
    show_gridlines: bool,
    font_family: str,
    font_size_pt: float,
) -> int:
    """چند ردیفِ *اول* در یک صفحه جا می‌شوند — با جستجویِ دودویی رویِ
    اندازه‌گیریِ واقعیِ QTextDocument.pageCount(). این فقط برایِ تعیینِ
    فاصله‌یِ درجِ ردیف‌هایِ «جمعِ این صفحه» استفاده می‌شود (نه برایِ شکستِ
    اجباریِ صفحه) — پس یک تخمینِ کمی نادرست هم صفحه‌یِ خالی نمی‌سازد،
    فقط ردیفِ جمع را یکی-دو ردیف زودتر/دیرتر می‌نشاند. طبقِ بازطراحیِ
    هدر/فوترِ تکرارشونده (که حالا بیرونِ محتوایِ جدول، به‌صورتِ نوارهایِ
    ثابت کشیده می‌شوند)، ظرفیتِ محتوایِ هر صفحه دیگر یکنواخت است — نیازی
    به تخمینِ جداگانه برایِ صفحه‌یِ اول نیست."""
    lo, hi, best = 1, len(rows), 1
    while lo <= hi:
        mid = (lo + hi) // 2
        trial_rows = rows[:mid]
        trial_extra_html = _bold_row_html(_subtotal_row(headers, amount_col_indices, trial_rows))
        table_html = _content_table_html(colgroup_html, head_cells, _rows_html(trial_rows), trial_extra_html, show_gridlines)
        if _fits_one_page(page_size, table_html, font_family, font_size_pt):
            best = mid
            lo = mid + 1
        else:
            hi = mid - 1
    return max(1, best)


def _rows_with_subtotals_html(
    rows: list[list],
    headers: list[str],
    amount_col_indices: list[int],
    rows_per_page_override: int,
    *,
    page_size: QSizeF,
    colgroup_html: str,
    head_cells: str,
    show_gridlines: bool,
    font_family: str,
    font_size_pt: float,
) -> str:
    if not amount_col_indices or len(rows) < 2:
        return _rows_html(rows)

    if rows_per_page_override and rows_per_page_override > 0:
        chunk_size = rows_per_page_override
    else:
        chunk_size = _estimate_rows_per_page(
            rows, headers, amount_col_indices, page_size=page_size, colgroup_html=colgroup_html,
            head_cells=head_cells, show_gridlines=show_gridlines, font_family=font_family, font_size_pt=font_size_pt,
        )

    parts = []
    remaining = rows
    while remaining:
        chunk = remaining[:chunk_size]
        parts.append(_rows_html(chunk))
        remaining = remaining[chunk_size:]
        if remaining:
            parts.append(_bold_row_html(_subtotal_row(headers, amount_col_indices, chunk)))
    return "".join(parts)


def _page_footer_html(footer_text: str, page_no: int, page_count: int) -> str:
    parts = []
    if footer_text.strip():
        parts.append(_multiline_html(footer_text))
    page_no_fa = numerals.to_persian_digits(str(page_no))
    page_count_fa = numerals.to_persian_digits(str(page_count))
    parts.append(
        f'<div style="text-align:center; font-size:8pt; color:#666;">صفحه‌یِ {page_no_fa} از {page_count_fa}</div>'
    )
    return "".join(parts)


def _apply_page_setup(printer: QPrinter) -> None:
    printer.setPageOrientation(QPageLayout.Orientation.Landscape)
    printer.setPageMargins(QMarginsF(_PAGE_MARGIN_MM, _PAGE_MARGIN_MM, _PAGE_MARGIN_MM, _PAGE_MARGIN_MM), QPageLayout.Unit.Millimeter)


def _paint_report(
    printer: QPrinter,
    title: str,
    headers: list[str],
    rows: list[list],
    footer: list | None,
    *,
    company_name: str,
    report_date: str,
    filters: list[tuple[str, str]] | None,
    options: PrintOptions | None = None,
) -> None:
    """طبقِ درخواستِ صریح («روی هر صفحه باید نشون بده صفحه چند از
    چنده و هدر و فوتر روی همه صفحات نمایش بده»): QTextDocument.print_
    به‌تنهایی هدر/فوترِ تکرارشونده یا شماره‌یِ صفحه نمی‌دهد — چاپ به‌صورتِ
    دستی، صفحه‌به‌صفحه، با QPainter انجام می‌شود: یک نوارِ هدر (ثابت،
    عیناً تکرارشونده) بالایِ هر صفحه، یک تکه از جدولِ محتوا در وسط، و یک
    نوارِ فوتر (متنِ کاربر + «صفحه‌یِ X از Y») پایینِ هر صفحه."""
    options = options or PrintOptions()
    font_family = _print_font_family()
    font_size_pt = options.font_size_pt
    header_html = _resolve_header_html(options, title, company_name, report_date, filters)
    reversed_headers = list(reversed(headers))
    head_cells = "".join(f"<th>{_escape(h)}</th>" for h in reversed_headers)
    colgroup_html = _colgroup_html(options.column_widths, len(headers))
    amount_col_indices = _amount_column_indices(headers, rows)
    footer_row_html = _bold_row_html(footer) if footer else ""

    # نکته‌یِ فنیِ مهم: QTextDocument همیشه محتوا را بر حسبِ Point
    # اندازه‌گیری می‌کند (نه device pixel چاپگر) — اگر page rect را
    # بر حسبِ DevicePixel (که رویِ چاپگرهای HighResolution عددی چند
    # برابرِ بزرگ‌تر است) به‌عنوانِ اندازه‌یِ صفحه بدهیم، کلِ گزارش
    # اشتباهاً «در یک صفحه‌یِ غول‌آسا» جا می‌شود (باگِ واقعیِ کشف‌شده حینِ
    # تست: گزارشِ ۱۶۰ردیفی هم فقط یک صفحه می‌شد). راه‌حل: همه‌ی
    # اندازه‌گیری‌ها/چیدمان بر حسبِ Point انجام می‌شود؛ painter هم با
    # همان نسبتِ device-pixel-به-point مقیاس می‌شود تا خروجیِ چاپیِ
    # واقعی درست بماند.
    page_rect = printer.pageRect(QPrinter.Unit.Point)
    page_rect_px = printer.pageRect(QPrinter.Unit.DevicePixel)
    page_width = page_rect.width()
    page_height = page_rect.height()

    header_doc = QTextDocument()
    header_doc.setHtml(_wrap_document(f'<div style="text-align:center;">{header_html}</div>', font_family, font_size_pt))
    header_doc.setTextWidth(page_width)
    header_height = header_doc.size().height()

    footer_probe_doc = QTextDocument()
    footer_probe_doc.setHtml(_wrap_document(_page_footer_html(options.footer_text, 1, 1), font_family, font_size_pt))
    footer_probe_doc.setTextWidth(page_width)
    footer_height = footer_probe_doc.size().height()

    content_height = max(50.0, page_height - header_height - footer_height)
    content_page_size = QSizeF(page_width, content_height)

    body_rows_html = _rows_with_subtotals_html(
        rows, headers, amount_col_indices, options.rows_per_page,
        page_size=content_page_size, colgroup_html=colgroup_html, head_cells=head_cells,
        show_gridlines=options.show_gridlines, font_family=font_family, font_size_pt=font_size_pt,
    )
    content_doc = QTextDocument()
    content_doc.setHtml(
        _wrap_document(
            _content_table_html(colgroup_html, head_cells, body_rows_html, footer_row_html, options.show_gridlines),
            font_family, font_size_pt,
        )
    )
    content_doc.setPageSize(content_page_size)
    page_count = max(1, content_doc.pageCount())

    painter = QPainter()
    if not painter.begin(printer):
        return
    try:
        if page_width and page_height:
            painter.scale(page_rect_px.width() / page_width, page_rect_px.height() / page_height)
        for page_index in range(page_count):
            painter.save()
            header_ctx = QAbstractTextDocumentLayout.PaintContext()
            header_ctx.clip = QRectF(0, 0, page_width, header_height)
            header_doc.documentLayout().draw(painter, header_ctx)
            painter.restore()

            painter.save()
            painter.translate(0, header_height)
            painter.setClipRect(QRectF(0, 0, page_width, content_height))
            painter.translate(0, -page_index * content_height)
            content_ctx = QAbstractTextDocumentLayout.PaintContext()
            content_ctx.clip = QRectF(0, page_index * content_height, page_width, content_height)
            content_doc.documentLayout().draw(painter, content_ctx)
            painter.restore()

            page_footer_doc = QTextDocument()
            page_footer_doc.setHtml(
                _wrap_document(_page_footer_html(options.footer_text, page_index + 1, page_count), font_family, font_size_pt)
            )
            page_footer_doc.setTextWidth(page_width)
            painter.save()
            painter.translate(0, header_height + content_height)
            footer_ctx = QAbstractTextDocumentLayout.PaintContext()
            footer_ctx.clip = QRectF(0, 0, page_width, footer_height)
            page_footer_doc.documentLayout().draw(painter, footer_ctx)
            painter.restore()

            if page_index < page_count - 1:
                printer.newPage()
    finally:
        painter.end()


class _PrintOptionsDialog(QDialog):
    def __init__(
        self, parent: QWidget | None, title: str, headers: list[str], defaults: PrintOptions | None,
        *, company_name: str, report_date: str, filters: list[tuple[str, str]] | None,
    ) -> None:
        super().__init__(parent)
        self.setWindowTitle(f"تنظیماتِ چاپ — {title}")
        self.setMinimumWidth(460)
        defaults = defaults or PrintOptions()
        layout = QVBoxLayout(self)

        top_row = QHBoxLayout()
        top_row.addWidget(QLabel("اندازه‌یِ فونت (pt):"))
        self.font_size_field = QSpinBox()
        self.font_size_field.setRange(6, 16)
        self.font_size_field.setValue(int(round(defaults.font_size_pt)))
        top_row.addWidget(self.font_size_field)
        top_row.addSpacing(16)
        top_row.addWidget(QLabel("ردیف در هر صفحه (۰=خودکار):"))
        self.rows_per_page_field = QSpinBox()
        self.rows_per_page_field.setRange(0, 500)
        self.rows_per_page_field.setValue(defaults.rows_per_page)
        top_row.addWidget(self.rows_per_page_field)
        top_row.addStretch(1)
        layout.addLayout(top_row)

        self.gridlines_checkbox = QCheckBox("نمایشِ خطوطِ جدول (ردیف‌بندی)")
        self.gridlines_checkbox.setChecked(defaults.show_gridlines)
        layout.addWidget(self.gridlines_checkbox)

        layout.addWidget(QLabel("عرضِ ستون‌ها (٪ از عرضِ کلِ جدول):"))
        columns_scroll = QScrollArea()
        columns_scroll.setWidgetResizable(True)
        columns_scroll.setMaximumHeight(200)
        columns_widget = QWidget()
        columns_form = QFormLayout(columns_widget)
        self._width_fields: list[QSpinBox] = []
        default_widths = defaults.column_widths if defaults.column_widths and len(defaults.column_widths) == len(headers) else None
        if default_widths is None:
            default_widths = [round(100 / max(len(headers), 1), 1)] * len(headers)
        for header, default_w in zip(headers, default_widths):
            field_widget = QSpinBox()
            field_widget.setRange(3, 60)
            field_widget.setValue(int(round(default_w)))
            field_widget.setSuffix(" ٪")
            columns_form.addRow(header, field_widget)
            self._width_fields.append(field_widget)
        columns_scroll.setWidget(columns_widget)
        layout.addWidget(columns_scroll)

        layout.addWidget(QLabel("متنِ هدر (کاملاً قابلِ‌ویرایش — همین متن به‌جایِ هدرِ خودکار چاپ می‌شود):"))
        self.header_text_field = QTextEdit()
        self.header_text_field.setPlainText(
            defaults.header_text or _default_header_plain_text(title, company_name, report_date, filters)
        )
        self.header_text_field.setMaximumHeight(90)
        layout.addWidget(self.header_text_field)

        layout.addWidget(QLabel("متنِ فوتر (اختیاری — رویِ همه‌ی صفحات همراه با «صفحه‌ی X از Y» تکرار می‌شود):"))
        self.footer_text_field = QTextEdit()
        self.footer_text_field.setPlainText(defaults.footer_text)
        self.footer_text_field.setMaximumHeight(70)
        layout.addWidget(self.footer_text_field)

        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.button(QDialogButtonBox.Ok).setText("تاییدِ تنظیمات و ادامه")
        buttons.button(QDialogButtonBox.Cancel).setText("انصراف")
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def result_options(self) -> PrintOptions:
        widths = [f.value() for f in self._width_fields]
        total = sum(widths) or 1
        normalized = [w * 100 / total for w in widths]
        return PrintOptions(
            font_size_pt=float(self.font_size_field.value()),
            column_widths=normalized,
            header_text=self.header_text_field.toPlainText().strip(),
            footer_text=self.footer_text_field.toPlainText().strip(),
            rows_per_page=self.rows_per_page_field.value(),
            show_gridlines=self.gridlines_checkbox.isChecked(),
        )


def prompt_print_options(
    parent_widget: QWidget,
    title: str,
    headers: list[str],
    defaults: PrintOptions | None = None,
    *,
    company_name: str = "",
    report_date: str = "",
    filters: list[tuple[str, str]] | None = None,
) -> PrintOptions | None:
    dialog = _PrintOptionsDialog(
        parent_widget, title, headers, defaults,
        company_name=company_name, report_date=report_date, filters=filters,
    )
    if dialog.exec() != QDialog.Accepted:
        return None
    return dialog.result_options()


def print_report(
    parent_widget: QWidget,
    title: str,
    headers: list[str],
    rows: list[list],
    footer: list | None = None,
    *,
    company_name: str = "",
    report_date: str = "",
    filters: list[tuple[str, str]] | None = None,
    options: PrintOptions | None = None,
) -> None:
    if not rows:
        QMessageBox.information(parent_widget, "چاپ", "گزارشی برایِ چاپ وجود ندارد.")
        return
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    _apply_page_setup(printer)
    preview = QPrintPreviewDialog(printer, parent_widget)
    preview.paintRequested.connect(
        lambda p: _paint_report(
            p, title, headers, rows, footer,
            company_name=company_name, report_date=report_date, filters=filters, options=options,
        )
    )
    # طبقِ درخواستِ صریح: فرمِ پیش‌نمایشِ چاپ به‌طورِ پیش‌فرض تمامِ صفحه باز شود.
    preview.showMaximized()
    preview.exec()


def export_report_pdf(
    parent_widget: QWidget,
    title: str,
    headers: list[str],
    rows: list[list],
    footer: list | None = None,
    *,
    company_name: str = "",
    report_date: str = "",
    filters: list[tuple[str, str]] | None = None,
    options: PrintOptions | None = None,
) -> None:
    if not rows:
        QMessageBox.information(parent_widget, "PDF", "گزارشی برایِ خروجیِ PDF وجود ندارد.")
        return
    path, _filter = QFileDialog.getSaveFileName(parent_widget, "ذخیره‌یِ PDF", f"{title}.pdf", "PDF (*.pdf)")
    if not path:
        return
    if not path.lower().endswith(".pdf"):
        path += ".pdf"
    printer = QPrinter(QPrinter.PrinterMode.HighResolution)
    printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
    _apply_page_setup(printer)
    printer.setOutputFileName(path)
    _paint_report(
        printer, title, headers, rows, footer,
        company_name=company_name, report_date=report_date, filters=filters, options=options,
    )
    QMessageBox.information(parent_widget, "خروجیِ PDF", "فایلِ PDF با موفقیت ساخته شد.")


def export_report_excel(
    parent_widget: QWidget,
    title: str,
    headers: list[str],
    rows: list[list],
    footer: list | None = None,
    *,
    company_name: str = "",
    report_date: str = "",
    filters: list[tuple[str, str]] | None = None,
    options: PrintOptions | None = None,
) -> None:
    if not rows:
        QMessageBox.information(parent_widget, "Excel", "گزارشی برایِ خروجیِ Excel وجود ندارد.")
        return
    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        QMessageBox.warning(parent_widget, "خطا", "امکانِ ساختِ فایلِ Excel روی این سیستم فراهم نیست.")
        return

    path, _filter = QFileDialog.getSaveFileName(parent_widget, "ذخیره‌یِ Excel", f"{title}.xlsx", "Excel (*.xlsx)")
    if not path:
        return
    if not path.lower().endswith(".xlsx"):
        path += ".xlsx"

    options = options or PrintOptions()
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = (title[:31] or "گزارش").replace("/", "-")
    sheet.sheet_view.rightToLeft = True

    # طبقِ همان درخواستِ سربرگِ چاپ: نامِ شرکت + عنوان + تاریخ/فیلترها (یا
    # متنِ سفارشیِ کاربر) به‌عنوانِ چند سطرِ اول، پیش از جدولِ خودِ گزارش.
    header_text = options.header_text.strip() or _default_header_plain_text(title, company_name, report_date, filters)
    for line in header_text.splitlines():
        sheet.append([line])
    sheet.cell(row=1, column=1).font = Font(bold=True, size=13)
    sheet.append([])

    header_row = sheet.max_row + 1
    sheet.append(headers)
    for cell in sheet[header_row]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append(list(row))
    if footer:
        sheet.append(list(footer))
        for cell in sheet[sheet.max_row]:
            cell.font = Font(bold=True)
    if options.footer_text.strip():
        sheet.append([])
        for line in options.footer_text.strip().splitlines():
            sheet.append([line])
    sheet.freeze_panes = f"A{header_row + 1}"

    if options.column_widths and len(options.column_widths) == len(headers):
        # درصدها را به واحدِ عرضِ اکسل (تقریباً کاراکتر) تبدیل می‌کنیم —
        # مجموعِ عرضِ همه‌ی ستون‌ها را حدودِ ۱۲۰ کاراکتر در نظر می‌گیریم.
        for col_index, width_pct in enumerate(options.column_widths, start=1):
            letter = sheet.cell(row=header_row, column=col_index).column_letter
            sheet.column_dimensions[letter].width = max(8, round(120 * width_pct / 100))
    else:
        for col_cells in sheet.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            sheet.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    workbook.save(path)
    QMessageBox.information(parent_widget, "خروجیِ Excel", "فایلِ Excel با موفقیت ساخته شد.")
